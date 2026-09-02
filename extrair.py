# -*- coding: utf-8 -*-
"""Le 'Roteiro Japao.xlsx' e produz dados.json com todo o conteudo do site."""
import openpyxl, json, datetime, re, os, sys

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(os.path.dirname(BASE), 'Roteiro Japão.xlsx')

def s(c):
    if c is None: return ''
    if isinstance(c, datetime.datetime): return c.strftime('%Y-%m-%d')
    if isinstance(c, float) and c == int(c): return str(int(c))
    return str(c).strip()

def num(c):
    try: return round(float(c), 2)
    except (TypeError, ValueError): return 0

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ---------- PASSEIOS: roteiro dia a dia ----------
ws = wb['Passeios']
rows = [[s(c) for c in r] for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=18, values_only=True)]

CIDADE = {'Osaka':'osaka','Kyoto':'kyoto','Kawaguchiko':'fuji','Tokyo':'tokyo','voo':'osaka'}
def cidade_de(base):
    alvo = base.split('→')[-1].strip() if '→' in base else base.split('/')[0].strip()
    return CIDADE.get(alvo, 'osaka')

dias, atual = [], None
for r in rows[5:103]:
    data, semana, base, periodo, ativ = r[0], r[1], r[2], r[3], r[4]
    if ativ == 'TOTAL DO DIA':
        if atual:
            atual['total'] = {'km': num(r[8]), 'min': num(r[9]), 'pe': num(r[10]),
                              'transp': num(r[11]), 'ingresso': num(r[12]), 'nota': r[17]}
        continue
    if not data: continue
    if not atual or atual['data'] != data:
        atual = {'data': data, 'semana': semana, 'base': base, 'cidade': cidade_de(base),
                 'transicao': '→' in base, 'atividades': [], 'total': {}}
        dias.append(atual)
    atual['atividades'].append({
        'periodo': periodo, 'nome': ativ, 'area': r[5], 'de': r[6], 'como': r[7],
        'km': num(r[8]), 'min': num(r[9]), 'pe': num(r[10]),
        'transp': num(r[11]), 'ingresso': num(r[12]),
        'horario': r[13], 'duracao': r[14], 'bebe': r[15], 'terreno': r[16], 'obs': r[17],
    })

# ---------- PASSEIOS: resumo geral ----------
resumo = [{'label': r[0], 'valor': num(r[4]), 'unidade': r[5], 'nota': r[7] if len(r) > 7 else ''}
          for r in rows[104:112] if r[0] and r[4]]

# ---------- PASSEIOS: conflitos, decisoes e alertas ----------
notas, pend = [], None
for r in rows[113:]:
    t = r[0]
    if not t: continue
    m = re.match(r'^(\d+)\.\s*(.+)$', t)
    if m:
        pend = {'n': int(m.group(1)), 'titulo': m.group(2), 'texto': ''}
        notas.append(pend)
    elif pend and not pend['texto']:
        pend['texto'] = t

# ---------- VOO ----------
wv = wb['Voo']
vr = [[s(c) for c in r] for r in wv.iter_rows(min_row=1, max_row=wv.max_row, max_col=10, values_only=True)]
voos = [{'numero': r[0], 'origem': r[1], 'destino': r[2], 'saidaData': r[3], 'saidaHora': r[4][:5] if r[4] else '',
         'chegadaData': r[5], 'chegadaHora': r[6][:5] if r[6] else '', 'duracao': r[7],
         'conexao': r[8], 'obs': r[9]} for r in vr[2:8] if r[0]]
voosNotas = []
bloco = None
for r in vr[9:]:
    t = r[0]
    if not t or r[1]: continue
    if len(t) < 70:                      # linha curta = titulo do bloco
        bloco = {'titulo': t, 'texto': ''}
        voosNotas.append(bloco)
    elif bloco:                          # linha longa = corpo do bloco atual
        bloco['texto'] = (bloco['texto'] + ' ' + t).strip()

# ---------- HOSPEDAGEM ----------
wh = wb['Hospedagem']
hr = [[s(c) for c in r] for r in wh.iter_rows(min_row=1, max_row=wh.max_row, max_col=12, values_only=True)]
def stays(ini, fim):
    out = []
    for r in hr[ini:fim]:
        if r[1] and r[3]:
            out.append({'hotel': r[1], 'link': r[2] if r[2].startswith('http') else '',
                        'cidade': r[3], 'checkin': r[4].replace('2026', '2027'),
                        'checkout': r[5].replace('2026', '2027'), 'noites': num(r[6]),
                        'total': num(r[10])})
    return out
hospedagem = {'a': stays(2, 26), 'b': stays(26, 38)}

dados = {'dias': dias, 'resumo': resumo, 'notas': notas, 'voos': voos,
         'voosNotas': voosNotas, 'hospedagem': hospedagem}

out = os.path.join(BASE, 'dados.json')
with open(out, 'w', encoding='utf-8') as f:
    json.dump(dados, f, ensure_ascii=False, indent=1)
print(f'OK  {len(dias)} dias, {sum(len(d["atividades"]) for d in dias)} atividades, '
      f'{len(notas)} notas, {len(voos)} voos, {len(hospedagem["a"])}+{len(hospedagem["b"])} hospedagens')
print('->', out, os.path.getsize(out), 'bytes')
