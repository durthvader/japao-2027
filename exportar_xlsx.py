# -*- coding: utf-8 -*-
"""Grava o roteiro final — com todos os ajustes aplicados — numa planilha nova.

    python exportar_xlsx.py

Sai em 'Roteiro Japao - FINAL.xlsx', na pasta acima, ao lado da planilha original.

Por que um arquivo separado e nao a planilha de sempre:

  * 'Roteiro Japao.xlsx' e a FONTE. O gerar_site.py reextrai sozinho quando ela fica
    mais nova que o dados.json; reescrever a fonte com o resultado dos ajustes faria
    o roteiro ja ajustado voltar como entrada e os ajustes serem aplicados de novo,
    por cima de si mesmos. Os pares de 'limpeza' tambem deixariam de casar, porque o
    texto cru que eles procuram nao existiria mais.
  * A planilha original tem a aba do acerto financeiro entre as familias. Este arquivo
    nasce do zero, so com o roteiro, entao pode ser mandado para o grupo inteiro.

Ou seja: a fonte continua sendo a planilha antiga mais o ajustes.json. Este arquivo e
uma fotografia do resultado, para ler, imprimir e compartilhar.
"""
import io, os, json, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import ajustes

BASE = os.path.dirname(os.path.abspath(__file__))
DESTINO = os.path.join(os.path.dirname(BASE), 'Roteiro Japao - FINAL.xlsx')

TINTA = {'osaka': 'FFE8D9', 'kyoto': 'E5E9F7', 'fuji': 'DEEDE4', 'tokyo': 'F3E2EE'}
ESCURO, MEIO, CLARO = '1F2933', '52606D', 'F5F7FA'
BORDA = Border(bottom=Side(style='thin', color='D8DEE6'))


def texto(v):
    return '' if v is None else str(v)


def cabecalho(ws, titulos, larguras):
    ws.append(titulos)
    for i, (t, w) in enumerate(zip(titulos, larguras), 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor=ESCURO)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'


def aba_roteiro(wb, djs):
    ws = wb.create_sheet('Roteiro')
    cols = ['Data', 'Semana', 'Base', 'Período', 'Atividade', 'Área', 'De', 'Como',
            'Km', 'Min', 'A pé', 'Transp ¥', 'Ingresso ¥', 'Horário', 'Duração',
            'Bebê', 'Terreno', 'Observações']
    cabecalho(ws, cols, [11, 10, 17, 8, 42, 15, 20, 34, 7, 7, 7, 10, 12, 22, 10, 14, 52, 90])

    for dia in djs['dias']:
        tinta = PatternFill('solid', fgColor=TINTA.get(dia['cidade'], CLARO))
        for k, a in enumerate(dia['atividades']):
            ws.append([dia['data'] if k == 0 else '', dia['semana'] if k == 0 else '',
                       dia['base'] if k == 0 else '',
                       texto(a.get('periodo')), texto(a.get('nome')), texto(a.get('area')),
                       texto(a.get('de')), texto(a.get('como')),
                       a.get('km') or 0, a.get('min') or 0, a.get('pe') or 0,
                       a.get('transp') or 0, a.get('ingresso') or 0,
                       texto(a.get('horario')), texto(a.get('duracao')),
                       texto(a.get('bebe')), texto(a.get('terreno')), texto(a.get('obs'))])
            linha = ws.max_row
            for i in range(1, len(cols) + 1):
                c = ws.cell(row=linha, column=i)
                c.alignment = Alignment(vertical='top', wrap_text=i in (5, 7, 8, 17, 18))
                c.border = BORDA
                if i <= 3:
                    c.fill = tinta
                    c.font = Font(bold=True, size=10)
            ws.cell(row=linha, column=5).font = Font(bold=True, size=10)

        t = dia.get('total', {})
        ws.append(['', '', '', '', 'TOTAL DO DIA', '', '', '',
                   t.get('km', 0), t.get('min', 0), t.get('pe', 0),
                   t.get('transp', 0), t.get('ingresso', 0), '', '', '', '',
                   texto(t.get('nota'))])
        linha = ws.max_row
        for i in range(1, len(cols) + 1):
            c = ws.cell(row=linha, column=i)
            c.fill = PatternFill('solid', fgColor=ESCURO)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.alignment = Alignment(vertical='center', wrap_text=i == 18)

    for linha in ws.iter_rows(min_row=2, min_col=9, max_col=13):
        for c in linha:
            c.number_format = '#,##0.#'
    return ws


def aba_mudancas(wb):
    """O historico das decisoes, com o motivo de cada uma — o que o site nao mostra."""
    ops = json.loads(io.open(os.path.join(BASE, 'ajustes.json'), encoding='utf-8').read())
    ws = wb.create_sheet('Mudanças')
    cabecalho(ws, ['#', 'Tipo', 'Dia', 'O quê', 'Por quê'], [5, 13, 22, 46, 110])
    for n, op in enumerate(ops['operacoes'], 1):
        tipo = op['tipo']
        if tipo == 'mover':
            dia = '%s → %s' % (op.get('de', '?'), op.get('para') or 'CORTADO')
        elif tipo == 'reatribuir':
            dia = ' / '.join('%s → %s' % kv for kv in op['troca'].items())
        else:
            dia = op.get('dia', '')
        alvo = op.get('nome') or (op.get('atividade') or {}).get('nome') or ''
        if tipo == 'nota':
            alvo = (op.get('texto', '')[:120] + '…') if op.get('texto') else ''
        ws.append([n, tipo, dia, alvo, op.get('motivo', '')])
        for i in range(1, 6):
            c = ws.cell(row=ws.max_row, column=i)
            c.alignment = Alignment(vertical='top', wrap_text=i >= 4)
            c.border = BORDA
    return len(ops['operacoes'])


def aba_simples(wb, nome, titulos, larguras, linhas):
    ws = wb.create_sheet(nome)
    cabecalho(ws, titulos, larguras)
    for r in linhas:
        ws.append(r)
        for i in range(1, len(titulos) + 1):
            c = ws.cell(row=ws.max_row, column=i)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = BORDA
    return ws


def main():
    dados = json.loads(io.open(os.path.join(BASE, 'dados.json'), encoding='utf-8').read())
    djs = ajustes.aplicar(dados)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    capa = wb.create_sheet('Leia-me')
    capa.column_dimensions['A'].width = 118
    hoje = datetime.date.today().strftime('%d/%m/%Y')
    for linha, estilo in [
        ('Japão 2027 — roteiro final', 'titulo'),
        ('16 a 31 de março de 2027 · Osaka, Kyoto, Kawaguchiko e Tokyo · 13 pessoas', 'sub'),
        ('', ''),
        ('Gerado em %s a partir da planilha original mais as decisões tomadas depois dela.' % hoje, ''),
        ('Esta é uma fotografia do resultado: a aba Roteiro já vem com tudo aplicado.', ''),
        ('', ''),
        ('Roteiro — os 16 dias, atividade por atividade, com total por dia.', ''),
        ('Mudanças — o que foi trocado em relação à planilha original, e por quê.', ''),
        ('Alertas — horários de fechamento, conflitos e o que precisa ser reservado.', ''),
        ('Voos e Hospedagem — como estão lançados.', ''),
        ('', ''),
        ('Não edite este arquivo esperando que o site mude: ele é a saída, não a entrada.', 'nota'),
        ('O site continua saindo de "Roteiro Japão.xlsx" + ajustes.json, na pasta site/.', 'nota'),
    ]:
        capa.append([linha])
        c = capa.cell(row=capa.max_row, column=1)
        c.alignment = Alignment(wrap_text=True, vertical='center')
        if estilo == 'titulo':
            c.font = Font(bold=True, size=20, color=ESCURO)
            capa.row_dimensions[capa.max_row].height = 30
        elif estilo == 'sub':
            c.font = Font(size=12, color=MEIO)
        elif estilo == 'nota':
            c.font = Font(size=10, italic=True, color=MEIO)

    ws = aba_roteiro(wb, djs)
    n_ops = aba_mudancas(wb)

    aba_simples(wb, 'Resumo', ['Item', 'Valor', 'Unidade', 'Nota'], [40, 12, 14, 90],
                [[r.get('label'), r.get('valor'), r.get('unidade'), r.get('nota')]
                 for r in djs.get('resumo', [])])

    aba_simples(wb, 'Alertas', ['#', 'Assunto', 'Detalhe'], [5, 42, 130],
                [[n.get('n'), n.get('titulo'), n.get('texto')]
                 for n in djs.get('notas', [])])

    aba_simples(wb, 'Voos', ['Voo', 'Origem', 'Destino', 'Saída', 'Hora', 'Chegada',
                             'Hora', 'Duração', 'Conexão', 'Observações'],
                [10, 16, 16, 12, 8, 12, 8, 11, 16, 70],
                [[v.get('numero'), v.get('origem'), v.get('destino'), v.get('saidaData'),
                  v.get('saidaHora'), v.get('chegadaData'), v.get('chegadaHora'),
                  v.get('duracao'), v.get('conexao'), v.get('obs')]
                 for v in djs.get('voos', [])])

    hosp = djs.get('hospedagem', {})
    aba_simples(wb, 'Hospedagem', ['Grupo', 'Hotel', 'Cidade', 'Check-in', 'Check-out',
                                   'Noites', 'Total R$'], [8, 40, 18, 12, 12, 8, 12],
                [[g.upper(), h.get('hotel'), h.get('cidade'), h.get('checkin'),
                  h.get('checkout'), h.get('noites'), h.get('total')]
                 for g in ('a', 'b') for h in hosp.get(g, [])])

    wb.save(DESTINO)
    n_ativ = sum(len(d['atividades']) for d in djs['dias'])
    print('OK  %d dias, %d atividades, %d decisões, %d linhas na aba Roteiro'
          % (len(djs['dias']), n_ativ, n_ops, ws.max_row - 1))
    print('->', DESTINO, os.path.getsize(DESTINO), 'bytes')


if __name__ == '__main__':
    main()
